import subprocess
import sys
import os
import urllib.request
import json
import time as systime
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pyautogui
import openpyxl
import winsound

# ตรวจสอบและติดตั้งไลบรารี
required_packages = ["tkinter", "pyautogui", "openpyxl", "pillow"]
def ensure_package(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    current_version = subprocess.check_output([sys.executable, "-m", "pip", "--version"]).decode()
    response = urllib.request.urlopen("https://pypi.org/pypi/pip/json")
    latest_version = json.load(response)["info"]["version"]
    if latest_version not in current_version:
        print(f"Updating pip to version {latest_version}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"])
except Exception as e:
    print(f"ไม่สามารถอัปเดต pip ได้: {e}")

for pkg in required_packages:
    ensure_package(pkg)

# ==== เริ่ม GUI ====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FOLDER = os.path.join(BASE_DIR, "ข้อมูลถาวร")
LOGIC_FOLDER = os.path.join(BASE_DIR, "Logic")
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(LOGIC_FOLDER, exist_ok=True)

logic_descriptions, available_logics = {}, []
for file in os.listdir(LOGIC_FOLDER):
    if file.endswith(".py"):
        name = file[:-3]
        with open(os.path.join(LOGIC_FOLDER, file), encoding="utf-8") as f:
            desc = f.readline().strip()
            logic_descriptions[name] = desc[1:].strip() if desc.startswith("#") else "ไม่มีคำอธิบาย"
        available_logics.append(name)

root = tk.Tk()
root.title("Lotus's Auto Typer V5 15-6-68")
root.geometry("512x512")

# ตัวแปรหลัก
data_list = []
typing_mode = tk.StringVar(value=available_logics[0] if available_logics else "")
delay_mode = tk.StringVar(value="ปรับเอง")
logic_description_var = tk.StringVar(value=logic_descriptions.get(typing_mode.get(), ""))
saved_file_var = tk.StringVar()
stop_typing_flag = False
delay_var = tk.DoubleVar(value=1.0)

# โหลด Logic
def load_typing_logic(mode):
    path = os.path.join(LOGIC_FOLDER, f"{mode}.py")
    namespace = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            exec(f.read(), namespace)
        return namespace.get("type_entry")
    except Exception as e:
        messagebox.showerror("โหลด Logic ล้มเหลว", str(e))
        return None

def load_saved_data(filename):
    filepath = os.path.join(DATA_FOLDER, filename)
    if not os.path.exists(filepath): return []
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    return [(str(row[0]), str(row[1]) if row[1] else "0") for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]

def refresh_display():
    display_box.config(state=tk.NORMAL)
    display_box.delete("1.0", tk.END)
    for i, (t, a) in enumerate(data_list, 1):
        display_box.insert(tk.END, f"{i}. {t} → {a}\n")
    display_box.config(state=tk.DISABLED)

def countdown(seconds):
    for i in range(seconds, 0, -1):
        progress_label.config(text=f"เริ่มพิมพ์ใน {i} วินาที...")
        winsound.Beep(1000, 500)
        root.update()
        systime.sleep(1)

def start_typing():
    global stop_typing_flag
    stop_typing_flag = False

    if not data_list:
        messagebox.showwarning("ไม่มีข้อมูล", "กรุณาอัปโหลดข้อมูลก่อน")
        return

    logic_func = load_typing_logic(typing_mode.get())
    if not logic_func:
        return

    def type_thread():
        countdown(3)

        start_idx = 0
        if os.path.exists("last_index.txt"):
            try:
                with open("last_index.txt", "r") as f:
                    start_idx = int(f.read().strip())
                print(f"📌 เริ่มจาก index {start_idx}")
            except:
                print("⚠️ อ่าน index ไม่ได้ เริ่มจาก 0")
                start_idx = 0

        total_items = len(data_list)

        if delay_mode.get() == "หน่วงแบบอัจฉริยะ":
            if total_items > 1000:
                auto_delay = 5.0
            elif total_items > 500:
                auto_delay = 3.0
            else:
                auto_delay = delay_var.get()
        else:
            auto_delay = delay_var.get()

        start_time = systime.time()
        for idx in range(start_idx, total_items):
            if stop_typing_flag:
                break

            text, amount = data_list[idx]
            logic_func(text, amount)
            pyautogui.press("enter")
            systime.sleep(auto_delay)

            with open("last_index.txt", "w") as f:
                f.write(str(idx + 1))

            progress = ((idx + 1) / total_items) * 100
            progress_var.set(progress)
            progress_label.config(text=f"{progress:.1f}% เสร็จแล้ว")

            display_box.config(state=tk.NORMAL)
            display_box.delete("1.0", tk.END)
            for i, (t, a) in enumerate(data_list, 1):
                mark = " ✓" if i <= idx + 1 else ""
                display_box.insert(tk.END, f"{i}. {t} → {a}{mark}\n")
            display_box.config(state=tk.DISABLED)

        if not stop_typing_flag and os.path.exists("last_index.txt"):
            os.remove("last_index.txt")

        elapsed = systime.time() - start_time
        mins, secs = divmod(elapsed, 60)
        messagebox.showinfo("เสร็จสิ้น", f"ใช้เวลา {int(mins)} นาที {secs:.2f} วินาที")

    threading.Thread(target=type_thread).start()

def stop_typing():
    global stop_typing_flag
    stop_typing_flag = True

def reset_index():
    if os.path.exists("last_index.txt"):
        os.remove("last_index.txt")
        print("🔁 รีเซ็ต index สำเร็จ")
    else:
        print("ℹ️ ไม่มีไฟล์ last_index.txt")
    messagebox.showinfo("รีเซ็ต", "จะเริ่มใหม่ตั้งแต่ต้นในการพิมพ์ครั้งหน้า")

def reset_and_restart():
    reset_index()
    messagebox.showinfo("รีเซ็ต", "กำลังรีโหลดโปรแกรมใหม่ทั้งหมด...")
    root.destroy()
    os.execl(sys.executable, sys.executable, *sys.argv)

def upload_file():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if path:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet = wb.active
            data_list.clear()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    data_list.append((str(row[0]), str(row[1]) if row[1] else "0"))
            refresh_display()
            messagebox.showinfo("อัปโหลด", f"โหลด {os.path.basename(path)} เรียบร้อย")
        except Exception as e:
            messagebox.showerror("โหลดล้มเหลว", str(e))

# ==== UI ====
frame_top = tk.Frame(root)
frame_top.pack(pady=10)
tk.Button(frame_top, text="📂 อัปโหลด Excel", command=upload_file, bg="#2196F3", fg="white").pack(side=tk.LEFT, padx=5)
tk.Button(frame_top, text="✅ เริ่มพิมพ์", command=start_typing, bg="green", fg="white").pack(side=tk.LEFT, padx=5)
tk.Button(frame_top, text="⛔ หยุดพิมพ์", command=stop_typing, bg="red", fg="white").pack(side=tk.LEFT, padx=5)
tk.Button(frame_top, text="🔁 เริ่มใหม่ทั้งหมด", command=reset_index, bg="#FFC107").pack(side=tk.LEFT, padx=5)
tk.Button(frame_top, text="🔄 รีเซ็ต + รีโหลด", command=reset_and_restart, bg="#FF5722", fg="white").pack(side=tk.LEFT, padx=5)

logic_frame = tk.Frame(root)
logic_frame.pack(pady=5)
tk.Label(logic_frame, text="🧠 เลือก Logic การพิมพ์:").pack(side=tk.LEFT)
logic_combo = ttk.Combobox(logic_frame, textvariable=typing_mode, values=available_logics, state="readonly")
logic_combo.pack(side=tk.LEFT, padx=5)
logic_combo.bind("<<ComboboxSelected>>", lambda e: logic_description_var.set(logic_descriptions.get(typing_mode.get(), "")))
logic_desc_label = tk.Label(root, textvariable=logic_description_var, fg="gray")
logic_desc_label.pack()

delay_frame = tk.Frame(root)
delay_frame.pack(pady=5)
tk.Label(delay_frame, text="⏱️ หน่วงเวลาต่อรายการ (วินาที):").pack(side=tk.LEFT)
tk.Entry(delay_frame, textvariable=delay_var, width=5).pack(side=tk.LEFT, padx=5)
tk.Label(delay_frame, text="โหมดหน่วง:").pack(side=tk.LEFT)
delay_combo = ttk.Combobox(delay_frame, textvariable=delay_mode, values=["ปรับเอง", "หน่วงแบบอัจฉริยะ"], state="readonly", width=20)
delay_combo.pack(side=tk.LEFT, padx=5)

text_frame = tk.Frame(root)
text_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
scrollbar = tk.Scrollbar(text_frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
display_box = tk.Text(text_frame, height=18, bg="#f0f0f0", yscrollcommand=scrollbar.set, state=tk.DISABLED)
display_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
scrollbar.config(command=display_box.yview)

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
progress_bar.pack(fill=tk.X, padx=20, pady=5)
progress_label = tk.Label(root, text="0.0%")
progress_label.pack()

# ==== ปุ่มปรับขนาดหน้าต่างที่มุมซ้ายล่าง ====
resize_frame = tk.Frame(root)
resize_frame.pack(side=tk.LEFT, anchor="sw", padx=10, pady=5)

tk.Label(resize_frame, text="🔧 ขนาดหน้าต่าง:").pack(side=tk.LEFT)
def resize_window(size_text):
    root.geometry(size_text)

resize_options = ["512x512", "680x640", "800x700", "1024x768"]
resize_combo = ttk.Combobox(resize_frame, values=resize_options, state="readonly", width=12)
resize_combo.set("เลือกขนาด")
resize_combo.pack(side=tk.LEFT, padx=5)
resize_combo.bind("<<ComboboxSelected>>", lambda e: resize_window(resize_combo.get()))

root.mainloop()
